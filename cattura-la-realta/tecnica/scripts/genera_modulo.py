#!/usr/bin/env python3
"""
genera_modulo.py
================

Genera Modulo_Sopralluogo_<id>.xlsx precompilato con i campi catastali letti
da dati_atto.md (prodotto dalla skill estrai-dati-atto-pregeo).

Esempio:
    python genera_modulo.py --dati-atto 03_OUTPUT/dati_atto.md --id 121 --out 02_RILIEVO/
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

DEFAULT_STRUMENTO = "Ricevitore GNSS marca e-Survey modello E300-Pro"
DEFAULT_TIPO_RILIEVO = "GPS RTK"


def parse_dati_atto(md_path: Path) -> dict:
    """Estrae Comune, Foglio, Particelle oggetto dal dati_atto.md."""
    out = {"comune": "", "foglio": "", "particelle": ""}
    if not md_path.exists():
        return out
    text = md_path.read_text(encoding="utf-8", errors="replace")

    def grab(label: str) -> str:
        m = re.search(rf"^- {re.escape(label)}:\s*(.+)$", text, re.MULTILINE)
        if not m:
            return ""
        v = m.group(1).strip()
        return "" if v.upper() in ("N.D.", "N.A.") else v

    out["comune"]     = grab("Comune")
    out["foglio"]     = grab("Foglio")
    out["particelle"] = grab("Particelle oggetto")
    return out


def costruisci_modulo(out_xlsx: Path, atto_id: str, comune: str,
                      foglio: str, particelle: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Serve openpyxl (pip install openpyxl).")

    wb = Workbook()

    # ----- Foglio DATI -----
    ws = wb.active
    ws.title = "DATI"
    ws["A1"] = f"Modulo Sopralluogo - Atto {atto_id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    ws.append([])
    ws.append(["Compila i campi vuoti. I campi catastali sono gia' precompilati "
               "dal dati_atto.md. Non modificare la colonna 'Campo'."])
    ws.append([])

    header_row = 5
    ws.cell(row=header_row, column=1, value="Campo").font = Font(bold=True)
    ws.cell(row=header_row, column=2, value="Valore").font = Font(bold=True)
    fill_h = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(row=header_row, column=1).fill = fill_h
    ws.cell(row=header_row, column=2).fill = fill_h

    rows = [
        ("Numero atto", atto_id),
        ("Comune", comune or ""),
        ("Foglio", foglio or ""),
        ("Particella oggetto", particelle or ""),
        ("Data sopralluogo", ""),
        ("Ora inizio", ""),
        ("Ora fine", ""),
        ("Strumentazione collaudo", DEFAULT_STRUMENTO),
        ("Tipo rilievo collaudo", DEFAULT_TIPO_RILIEVO),
        ("Esito sintetico", ""),
        ("Note generali", ""),
    ]
    side = Side(style="thin", color="CCCCCC")
    border = Border(top=side, bottom=side, left=side, right=side)
    fill_pre     = PatternFill("solid", fgColor="E8F4EA")  # verde chiaro per precompilati
    fill_default = PatternFill("solid", fgColor="FFF7D6")  # giallo chiaro per default modificabili

    precompilati = {"Numero atto", "Comune", "Foglio", "Particella oggetto"}
    con_default  = {"Strumentazione collaudo", "Tipo rilievo collaudo"}

    for i, (campo, valore) in enumerate(rows, start=header_row + 1):
        c1 = ws.cell(row=i, column=1, value=campo)
        c2 = ws.cell(row=i, column=2, value=valore)
        c1.font = Font(bold=True)
        c1.border = border
        c2.border = border
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        if campo in precompilati:
            c2.fill = fill_pre
        elif campo in con_default:
            c2.fill = fill_default

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.row_dimensions[header_row + 11].height = 60  # Note generali

    # Legenda colori in fondo
    ws.cell(row=header_row + 13, column=1, value="Legenda:").font = Font(italic=True, bold=True)
    ws.cell(row=header_row + 14, column=1, value="verde = gia' precompilato dal dati_atto.md").font = Font(italic=True)
    ws.cell(row=header_row + 15, column=1, value="giallo = default modificabile").font = Font(italic=True)
    ws.cell(row=header_row + 16, column=1, value="bianco = da compilare").font = Font(italic=True)

    # ----- Foglio PARTECIPANTI -----
    ws2 = wb.create_sheet("PARTECIPANTI")
    ws2.append(["Cognome e nome", "Ruolo", "Organizzazione"])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = fill_h
        c.border = border
    # 8 righe vuote pronte
    for _ in range(8):
        ws2.append(["", "", ""])
        for c in ws2[ws2.max_row]:
            c.border = border
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 30

    # ----- Foglio EVIDENZE -----
    ws3 = wb.create_sheet("EVIDENZE")
    ws3.append(["Argomento", "Principali evidenze e azione conseguente"])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = fill_h
        c.border = border
    # 12 righe vuote pronte (con altezza maggiore)
    for _ in range(12):
        ws3.append(["", ""])
        for c in ws3[ws3.max_row]:
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[ws3.max_row].height = 50
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 80

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Genera Modulo_Sopralluogo precompilato.")
    ap.add_argument("--dati-atto", type=Path, default=None,
                    help="Percorso a dati_atto.md (per i campi catastali). "
                         "Se omesso, i campi catastali restano vuoti.")
    ap.add_argument("--id", required=True, help="Numero atto (es. 121).")
    ap.add_argument("--out", type=Path, default=Path("."),
                    help="Cartella di output (tipicamente 02_RILIEVO/).")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args(argv)

    dati = {"comune": "", "foglio": "", "particelle": ""}
    if args.dati_atto:
        dati = parse_dati_atto(args.dati_atto)

    out_path = args.out / f"Modulo_Sopralluogo_{args.id}.xlsx"
    costruisci_modulo(out_path, args.id, dati["comune"], dati["foglio"], dati["particelle"])

    if not args.quiet:
        print(f"Modulo generato: {out_path}")
        if any(dati.values()):
            print(f"Precompilato con: Comune={dati['comune']}, Foglio={dati['foglio']}, Particelle={dati['particelle']}")
        else:
            print("Nessun campo catastale precompilato (manca dati_atto.md).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
