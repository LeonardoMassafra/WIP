#!/usr/bin/env python3
"""
compare_pairs.py
================

Confronta le mutue distanze fra coppie omologhe di punti partendo da due file
di coordinate (CSV/TXT/XLSX). Tipico uso: collaudo di un atto di aggiornamento
catastale dove il file A è il rilievo del frazionamento approvato e il file B
è il rilievo di collaudo dello studio.

Output:
    - <out_dir>/Confronto_misurate_<job>.xlsx (4 fogli: Coppie, Calcoli, Sintesi, Grafici)
    - <out_dir>/confronto_misurate.md
    - <out_dir>/grafici/delta_per_coppia.png
    - <out_dir>/grafici/mappa_punti.png
    - <out_dir>/grafici/conformita_per_tipo.png

L'Excel è generato da `_excel.py` e i PNG da `_viz.py` (matplotlib opzionale).

Esempio d'uso:

    python compare_pairs.py \\
        --frazionamento rilievi/fraz.csv \\
        --collaudo rilievi/coll.csv \\
        --job 25022-CR-001 \\
        --comune Bovolone \\
        --foglio 14 \\
        --mappali "123, 124, 125" \\
        --data-collaudo 22/04/2026 \\
        --out outputs/

Tolleranze di default: PF-PF 0.20 m, Det-Det 0.05 m, mista 0.20 m.
Cambiabili con --tol-pf, --tol-det, --tol-mista.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from itertools import combinations
from pathlib import Path

PF_PATTERN = re.compile(r"^\s*PF\s*\d+\s*/\s*\d+\s*/\s*[A-Z]\d+\s*$", re.IGNORECASE)

COL_ALIASES = {
    "nome":  {"nome", "name", "punto", "point", "id", "pt", "p"},
    "est":   {"est", "e", "x"},
    "nord":  {"nord", "n", "y"},
    "quota": {"quota", "q", "z", "h"},
}


@dataclass
class Punto:
    nome: str
    est: float
    nord: float
    quota: float | None
    is_pf: bool = False


@dataclass
class Coppia:
    nome_a: str
    nome_b: str
    tipo: str
    dist_fraz: float
    dist_coll: float
    delta: float
    soglia: float
    esito: str


def _detect_dialect(sample: str) -> tuple[str | None, str]:
    lines = [l for l in sample.splitlines() if l.strip()]
    if not lines:
        raise ValueError("File vuoto")
    test = lines[1] if len(lines) > 1 else lines[0]
    counts = {sep: test.count(sep) for sep in [";", ",", "\t", "|"]}
    sep_col = max(counts, key=counts.get)
    if counts[sep_col] == 0:
        sep_col = None
    has_comma_dec = bool(re.search(r"\d,\d", test))
    sep_dec = "." if (sep_col == "," or not has_comma_dec) else ","
    return sep_col, sep_dec


def _split_row(row: str, sep_col: str | None) -> list[str]:
    if sep_col is None:
        return row.split()
    if sep_col == "\t":
        return row.split("\t")
    return [c.strip() for c in row.split(sep_col)]


def _parse_float(s: str, sep_dec: str) -> float:
    s = s.strip().replace(" ", "")
    if sep_dec == ",":
        s = s.replace(".", "").replace(",", ".")
    return float(s)


def _looks_like_header(row: list[str]) -> bool:
    lowered = [c.strip().lower() for c in row]
    matches = sum(1 for c in lowered if any(c in a for a in COL_ALIASES.values()))
    return matches >= 3


def _resolve_column_order(header: list[str]) -> list[int | None]:
    lowered = [c.strip().lower() for c in header]
    order: list[int | None] = []
    for key in ("nome", "est", "nord", "quota"):
        idx = next((i for i, c in enumerate(lowered) if c in COL_ALIASES[key]), None)
        if idx is None and key != "quota":
            raise ValueError(f"Header senza colonna '{key}': {header}")
        order.append(idx)
    return order


def _read_xlsx(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Per leggere XLSX serve openpyxl (pip install openpyxl).")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    return [[("" if c is None else str(c)) for c in row]
            for row in ws.iter_rows(values_only=True)]


def read_coords(path: Path, pf_list: set[str] | None = None) -> list[Punto]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        rows_raw = _read_xlsx(path)
        rows = [[c.strip() for c in r] for r in rows_raw if any(str(c).strip() for c in r)]
        sep_dec = "."
    else:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        sep_col, sep_dec = _detect_dialect(text)
        rows = []
        for line in text.splitlines():
            if not line.strip():
                continue
            rows.append(_split_row(line, sep_col))

    if not rows:
        raise ValueError(f"File senza dati: {path}")

    has_header = _looks_like_header(rows[0])
    if has_header:
        order = _resolve_column_order(rows[0])
        data_rows = rows[1:]
    else:
        order = [0, 1, 2, 3 if len(rows[0]) >= 4 else None]
        data_rows = rows

    points: list[Punto] = []
    pf_norm = {p.upper().replace(" ", "") for p in (pf_list or [])}

    for r in data_rows:
        if not r or all(not c for c in r):
            continue
        try:
            nome = r[order[0]].strip()
            if not nome:
                continue
            est = _parse_float(r[order[1]], sep_dec)
            nord = _parse_float(r[order[2]], sep_dec)
            quota = None
            if order[3] is not None and len(r) > order[3] and r[order[3]].strip():
                try:
                    quota = _parse_float(r[order[3]], sep_dec)
                except ValueError:
                    quota = None
        except (IndexError, ValueError) as e:
            print(f"[avviso] riga ignorata in {path.name}: {r} ({e})", file=sys.stderr)
            continue
        is_pf = bool(PF_PATTERN.match(nome))
        if not is_pf and pf_norm and nome.upper().replace(" ", "") in pf_norm:
            is_pf = True
        points.append(Punto(nome=nome, est=est, nord=nord, quota=quota, is_pf=is_pf))

    return points


def distance(a: Punto, b: Punto, three_d: bool = False) -> float:
    de, dn = a.est - b.est, a.nord - b.nord
    if three_d:
        if a.quota is None or b.quota is None:
            raise ValueError(f"Distanza 3D ma manca quota su {a.nome} o {b.nome}")
        dz = a.quota - b.quota
        return (de*de + dn*dn + dz*dz) ** 0.5
    return (de*de + dn*dn) ** 0.5


def classifica_tipo(p: Punto, q: Punto) -> str:
    if p.is_pf and q.is_pf:
        return "PF-PF"
    if p.is_pf or q.is_pf:
        return "PF-Det"
    return "Det-Det"


def confronta(fraz, coll, tol_pf=0.20, tol_det=0.05, tol_mista=0.20, three_d=False):
    nomi_f = {p.nome.upper().strip(): p for p in fraz}
    nomi_c = {p.nome.upper().strip(): p for p in coll}
    omologhi = sorted(set(nomi_f) & set(nomi_c))
    solo_f = sorted(set(nomi_f) - set(nomi_c))
    solo_c = sorted(set(nomi_c) - set(nomi_f))

    coppie: list[Coppia] = []
    for a, b in combinations(omologhi, 2):
        pa_f, pb_f = nomi_f[a], nomi_f[b]
        pa_c, pb_c = nomi_c[a], nomi_c[b]
        tipo = classifica_tipo(pa_f, pb_f)
        soglia = {"PF-PF": tol_pf, "Det-Det": tol_det, "PF-Det": tol_mista}[tipo]
        d_f = distance(pa_f, pb_f, three_d)
        d_c = distance(pa_c, pb_c, three_d)
        delta = abs(d_f - d_c)
        esito = "OK" if delta <= soglia else "NOK"
        coppie.append(Coppia(pa_f.nome, pb_f.nome, tipo, d_f, d_c, delta, soglia, esito))

    map_f = {p.nome.upper().strip(): p.nome for p in fraz}
    map_c = {p.nome.upper().strip(): p.nome for p in coll}
    return coppie, [map_f[k] for k in solo_f], [map_c[k] for k in solo_c]


# --------------------------------------------------------------------------- #
# Markdown
# --------------------------------------------------------------------------- #

def fmt(v: float, n: int = 3) -> str:
    return f"{v:.{n}f}"


def scrivi_markdown(coppie, out_path, job, comune, foglio, mappali, data_collaudo,
                    file_fraz, file_coll, solo_fraz, solo_coll,
                    tol_pf, tol_det, tol_mista, note, grafici_disponibili=None):
    L: list[str] = []
    L.append(f"# Confronto misurate - {job}")
    L.append("")
    L.append("## Identificazione lavoro")
    L.append(f"- Commessa: {job}")
    L.append(f"- Comune: {comune or 'N.D.'}")
    L.append(f"- Foglio: {foglio or 'N.D.'}")
    L.append(f"- Mappali oggetto: {mappali or 'N.D.'}")
    L.append(f"- Data collaudo: {data_collaudo or 'N.D.'}")
    L.append(f"- File frazionamento: {file_fraz}")
    L.append(f"- File collaudo: {file_coll}")
    L.append("")
    L.append("## Tolleranze applicate")
    L.append(f"- PF - PF: {fmt(tol_pf, 2)} m")
    L.append(f"- Dettaglio - Dettaglio: {fmt(tol_det, 2)} m")
    L.append(f"- PF - Dettaglio (mista): {fmt(tol_mista, 2)} m")
    L.append("")

    L.append("## Sintesi")
    L.append("| Tipo coppia | Coppie totali | OK | NOK | % conformita | max |D| (m) |")
    L.append("|---|---|---|---|---|---|")
    label_map = {"PF-PF": "PF - PF", "PF-Det": "PF - Dettaglio",
                 "Det-Det": "Dettaglio - Dettaglio"}
    for tipo in ("PF-PF", "PF-Det", "Det-Det"):
        sub = [c for c in coppie if c.tipo == tipo]
        if not sub:
            continue
        ok = sum(1 for c in sub if c.esito == "OK")
        nok = len(sub) - ok
        pct = ok / len(sub) * 100
        mx = max(c.delta for c in sub)
        L.append(f"| {label_map[tipo]} | {len(sub)} | {ok} | {nok} | {pct:.1f}% | {fmt(mx)} |")
    if coppie:
        ok = sum(1 for c in coppie if c.esito == "OK")
        nok = len(coppie) - ok
        pct = ok / len(coppie) * 100
        mx = max(c.delta for c in coppie)
        L.append(f"| **Totale** | {len(coppie)} | {ok} | {nok} | **{pct:.1f}%** | {fmt(mx)} |")
    L.append("")

    sezioni = [("PF-PF",   "Coppie PF - PF"),
               ("PF-Det",  "Coppie PF - Dettaglio"),
               ("Det-Det", "Coppie Dettaglio - Dettaglio")]
    for tipo, titolo in sezioni:
        sub = [c for c in coppie if c.tipo == tipo]
        if not sub:
            continue
        L.append(f"## {titolo}")
        L.append("| # | Punto A | Punto B | Dist. frazionamento (m) | Dist. collaudo (m) | |D| (m) | Soglia (m) | Esito |")
        L.append("|---|---|---|---|---|---|---|---|")
        for i, c in enumerate(sorted(sub, key=lambda x: -x.delta), start=1):
            L.append(f"| {i} | {c.nome_a} | {c.nome_b} | {fmt(c.dist_fraz)} | "
                     f"{fmt(c.dist_coll)} | {fmt(c.delta)} | {fmt(c.soglia, 2)} | {c.esito} |")
        L.append("")

    L.append("## Punti senza omologo")
    L.append(f"- Presenti solo nel frazionamento: {', '.join(solo_fraz) if solo_fraz else 'nessuno'}")
    L.append(f"- Presenti solo nel collaudo: {', '.join(solo_coll) if solo_coll else 'nessuno'}")
    L.append("")

    if grafici_disponibili:
        L.append("## Grafici disponibili")
        for p in grafici_disponibili:
            L.append(f"- `{p.name}` (cartella `grafici/`)")
        L.append("")

    L.append("## Note")
    if note:
        for n in note:
            L.append(f"- {n}")
    else:
        L.append("- Nessuna nota.")
    L.append("")

    out_path.write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Confronto delle mutue distanze fra coppie omologhe di punti.",
    )
    ap.add_argument("--frazionamento", required=True, type=Path)
    ap.add_argument("--collaudo", required=True, type=Path)
    ap.add_argument("--job", default="lavoro")
    ap.add_argument("--comune", default=None)
    ap.add_argument("--foglio", default=None)
    ap.add_argument("--mappali", default=None)
    ap.add_argument("--data-collaudo", default=None)
    ap.add_argument("--out", type=Path, default=Path("."))
    ap.add_argument("--tol-pf",    type=float, default=0.20)
    ap.add_argument("--tol-det",   type=float, default=0.05)
    ap.add_argument("--tol-mista", type=float, default=0.20)
    ap.add_argument("--3d", dest="three_d", action="store_true")
    ap.add_argument("--pf-list", default=None)
    ap.add_argument("--no-grafici", action="store_true",
                    help="Salta la generazione dei PNG. L'Excel ha comunque i grafici nativi.")
    ap.add_argument("--quiet", action="store_true")

    args = ap.parse_args(argv)
    args.out.mkdir(parents=True, exist_ok=True)

    pf_list = None
    if args.pf_list:
        pf_list = {p.strip() for p in args.pf_list.split(",") if p.strip()}

    note: list[str] = []
    fraz = read_coords(args.frazionamento, pf_list=pf_list)
    coll = read_coords(args.collaudo, pf_list=pf_list)
    if not fraz: sys.exit(f"Nessun punto valido in {args.frazionamento}")
    if not coll: sys.exit(f"Nessun punto valido in {args.collaudo}")

    if not any(p.is_pf for p in fraz) and not any(p.is_pf for p in coll):
        note.append("Nessun PF riconosciuto (pattern PFnn/nnnn/Xnnn). "
                    "Se i nomi PF sono custom, rilancia con --pf-list.")

    coppie, solo_fraz, solo_coll = confronta(
        fraz, coll, tol_pf=args.tol_pf, tol_det=args.tol_det,
        tol_mista=args.tol_mista, three_d=args.three_d)

    if not coppie:
        sys.exit("Nessuna coppia omologa: i due file non condividono nomi punto.")

    if solo_fraz or solo_coll:
        note.append(f"Punti senza omologo esclusi: "
                    f"{len(solo_fraz)} solo frazionamento, {len(solo_coll)} solo collaudo.")
    if args.three_d:
        note.append("Distanze calcolate in 3D (Est, Nord, Quota).")
    if (args.tol_pf, args.tol_det, args.tol_mista) != (0.20, 0.05, 0.20):
        note.append(f"Tolleranze custom: PF {args.tol_pf:.2f} m, "
                    f"Det {args.tol_det:.2f} m, mista {args.tol_mista:.2f} m.")

    job_safe = re.sub(r"[^A-Za-z0-9_\-]+", "_", args.job)
    xlsx_path = args.out / f"Confronto_misurate_{job_safe}.xlsx"
    md_path   = args.out / "confronto_misurate.md"

    fraz_dict = {p.nome.upper().strip(): p for p in fraz}
    coll_dict = {p.nome.upper().strip(): p for p in coll}

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from _excel import scrivi_excel
    scrivi_excel(coppie, xlsx_path, args.job, args.comune, args.foglio,
                 solo_fraz, solo_coll, args.tol_pf, args.tol_det, args.tol_mista,
                 fraz_dict, coll_dict)

    png_files: list[Path] = []
    if not args.no_grafici:
        try:
            from _viz import genera_grafici_png
            png_files = genera_grafici_png(
                coppie, fraz, coll, args.out,
                tol_pf=args.tol_pf, tol_det=args.tol_det, tol_mista=args.tol_mista)
        except Exception as e:
            print(f"[avviso] PNG non generati: {e}", file=sys.stderr)

    scrivi_markdown(coppie, md_path, args.job, args.comune, args.foglio,
                    args.mappali, args.data_collaudo,
                    args.frazionamento.name, args.collaudo.name,
                    solo_fraz, solo_coll,
                    args.tol_pf, args.tol_det, args.tol_mista,
                    note, grafici_disponibili=png_files)

    if not args.quiet:
        ok = sum(1 for c in coppie if c.esito == "OK")
        print(f"OK: {ok}/{len(coppie)} coppie ({ok/len(coppie)*100:.1f}% conformita). "
              f"NOK: {len(coppie)-ok}.")
        print(f"Excel: {xlsx_path}")
        print(f"Markdown: {md_path}")
        for p in png_files:
            print(f"Grafico: {p}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
