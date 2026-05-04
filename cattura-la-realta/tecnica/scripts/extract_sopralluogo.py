#!/usr/bin/env python3
"""
extract_sopralluogo.py
======================

Legge un Modulo_Sopralluogo_<id>.xlsx compilato a mano dopo la campagna
e produce dati_sopralluogo.md strutturato. Stesso pattern di
estrai-dati-per-accesso-agli-atti -> dati_estratti.md.

Esempio:
    python extract_sopralluogo.py --modulo Modulo_Sopralluogo_121.xlsx --out 03_OUTPUT/
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path


def leggi_modulo(xlsx_path: Path) -> dict:
    """Legge i 3 fogli e restituisce un dict strutturato."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Serve openpyxl (pip install openpyxl).")

    wb = load_workbook(xlsx_path, data_only=True)

    out: dict = {"dati": {}, "partecipanti": [], "evidenze": []}

    # ----- DATI -----
    if "DATI" in wb.sheetnames:
        ws = wb["DATI"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0] and isinstance(row[0], str):
                campo = str(row[0]).strip()
                # Salta header, titolo, legenda e righe spiegazione
                if (campo.lower() in ("campo", "modulo sopralluogo", "legenda:", "compila i campi vuoti.")
                        or campo.startswith("verde") or campo.startswith("giallo")
                        or campo.startswith("bianco") or campo.startswith("Modulo Sopralluogo")
                        or campo.startswith("Compila i campi")):
                    continue
                valore = row[1] if len(row) > 1 else None
                if valore is None or (isinstance(valore, str) and not valore.strip()):
                    out["dati"][campo] = ""
                else:
                    if isinstance(valore, datetime):
                        # data o time
                        if valore.hour or valore.minute:
                            out["dati"][campo] = valore.strftime("%H:%M")
                        else:
                            out["dati"][campo] = valore.strftime("%d/%m/%Y")
                    else:
                        out["dati"][campo] = str(valore).strip()

    # ----- PARTECIPANTI -----
    if "PARTECIPANTI" in wb.sheetnames:
        ws = wb["PARTECIPANTI"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            cog = (row[0] or "").strip() if isinstance(row[0], str) else (str(row[0]) if row[0] else "")
            ruolo = (row[1] or "").strip() if isinstance(row[1], str) else (str(row[1]) if row[1] else "")
            org = (row[2] or "").strip() if isinstance(row[2], str) and len(row) > 2 else (str(row[2]) if len(row) > 2 and row[2] else "")
            if cog or ruolo or org:
                out["partecipanti"].append({"nome": cog, "ruolo": ruolo, "organizzazione": org})

    # ----- EVIDENZE -----
    if "EVIDENZE" in wb.sheetnames:
        ws = wb["EVIDENZE"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            arg = (row[0] or "").strip() if isinstance(row[0], str) else (str(row[0]) if row[0] else "")
            ev = (row[1] or "").strip() if isinstance(row[1], str) and len(row) > 1 else (str(row[1]) if len(row) > 1 and row[1] else "")
            if arg or ev:
                out["evidenze"].append({"argomento": arg, "evidenze": ev})

    return out


def to_markdown(data: dict, atto_id: str, modulo_name: str) -> str:
    L: list[str] = []
    A = L.append
    d = data["dati"]

    def v(key: str) -> str:
        val = d.get(key, "").strip()
        return val if val else "N.D."

    A(f"# Sopralluogo - {atto_id}")
    A("")
    A("## Identificazione")
    A(f"- Numero atto: {v('Numero atto') or atto_id}")
    A(f"- Comune: {v('Comune')}")
    A(f"- Foglio: {v('Foglio')}")
    A(f"- Particella oggetto: {v('Particella oggetto')}")
    A("")
    A("## Sopralluogo")
    A(f"- Data: {v('Data sopralluogo')}")
    A(f"- Ora inizio: {v('Ora inizio')}")
    A(f"- Ora fine: {v('Ora fine')}")
    A(f"- Esito sintetico: {v('Esito sintetico')}")
    A("")
    A("## Strumentazione e rilievo")
    A(f"- Strumentazione: {v('Strumentazione collaudo')}")
    A(f"- Tipo rilievo: {v('Tipo rilievo collaudo')}")
    A("")
    A("## Partecipanti")
    if data["partecipanti"]:
        for p in data["partecipanti"]:
            parts = []
            if p["nome"]:          parts.append(p["nome"])
            extra = []
            if p["ruolo"]:         extra.append(p["ruolo"])
            if p["organizzazione"]: extra.append(p["organizzazione"])
            line = parts[0] if parts else ""
            if extra:
                line = f"{line} ({', '.join(extra)})" if line else f"({', '.join(extra)})"
            A(f"- {line}")
    else:
        A("- N.D.")
    A("")
    A("## Evidenze emerse")
    if data["evidenze"]:
        for ev in data["evidenze"]:
            arg = ev["argomento"] or "(senza titolo)"
            A(f"### {arg}")
            A(ev["evidenze"] or "N.D.")
            A("")
    else:
        A("- Nessuna evidenza registrata.")
        A("")
    A("## Note generali")
    note = v("Note generali")
    A(note if note != "N.D." else "- N.D.")
    A("")
    A("## File di origine")
    A(f"- Modulo: {modulo_name}")
    A(f"- Data estrazione: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    A("")
    return "\n".join(L)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Estrai dati dal Modulo_Sopralluogo compilato.")
    ap.add_argument("--modulo", required=True, type=Path,
                    help="Percorso al Modulo_Sopralluogo_<id>.xlsx compilato.")
    ap.add_argument("--id", default=None, help="Numero atto (default: dal nome file).")
    ap.add_argument("--out", type=Path, default=Path("."), help="Cartella di output.")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args(argv)

    args.out.mkdir(parents=True, exist_ok=True)

    # Atto id dal nome file: Modulo_Sopralluogo_121.xlsx -> 121
    atto_id = args.id
    if not atto_id:
        import re as _re
        m = _re.search(r"Modulo_Sopralluogo_(\S+?)\.xlsx", args.modulo.name)
        atto_id = m.group(1) if m else args.modulo.stem

    data = leggi_modulo(args.modulo)
    md = to_markdown(data, atto_id, args.modulo.name)

    out_path = args.out / "dati_sopralluogo.md"
    out_path.write_text(md, encoding="utf-8")

    if not args.quiet:
        print(f"Sopralluogo {atto_id}: {len(data['partecipanti'])} partecipanti, "
              f"{len(data['evidenze'])} evidenze.")
        nd = sum(1 for v in data["dati"].values() if not v)
        if nd:
            print(f"Campi vuoti (saranno N.D. nel report): {nd}")
        print(f"Output: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
