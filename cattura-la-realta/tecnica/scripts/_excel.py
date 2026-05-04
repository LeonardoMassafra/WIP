"""Generatore dell'Excel di lavoro per compare_pairs.py.

Quattro fogli: Coppie (riepilogo), Calcoli (formule live), Sintesi, Grafici.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from compare_pairs import Coppia, Punto


def scrivi_excel(coppie, out_path: Path, job: str,
                 comune: str | None, foglio: str | None,
                 solo_fraz: list[str], solo_coll: list[str],
                 tol_pf: float, tol_det: float, tol_mista: float,
                 fraz_dict, coll_dict) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Serve openpyxl (pip install openpyxl).")

    wb = Workbook()

    # ----- Foglio Coppie -----
    ws = wb.active
    ws.title = "Coppie"
    ws.append(["Tipo coppia", "Punto A", "Punto B",
               "Dist. frazionamento (m)", "Dist. collaudo (m)",
               "|D| (m)", "Soglia (m)", "Esito"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    fill_nok = PatternFill("solid", fgColor="FFC7CE")
    fill_ok  = PatternFill("solid", fgColor="C6EFCE")
    ordine = {"PF-PF": 0, "PF-Det": 1, "Det-Det": 2}
    coppie_ord = sorted(coppie, key=lambda c: (ordine[c.tipo], -c.delta))
    for c in coppie_ord:
        ws.append([c.tipo, c.nome_a, c.nome_b,
                   round(c.dist_fraz, 3), round(c.dist_coll, 3),
                   round(c.delta, 3), round(c.soglia, 3), c.esito])
        fill = fill_nok if c.esito == "NOK" else fill_ok
        for cell in ws[ws.max_row]:
            cell.fill = fill
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    for i, w in enumerate([14, 18, 18, 22, 22, 12, 12, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ----- Foglio Calcoli (formule live) -----
    ws_calc = wb.create_sheet("Calcoli")
    ws_calc.append([f"Dimostrazione del calcolo - {job}"])
    ws_calc["A1"].font = Font(bold=True, size=14)
    ws_calc.append([])
    ws_calc.append(["Le distanze e i delta sono FORMULE LIVE: cambia una "
                    "coordinata e tutto si ricalcola. dist = SQRT(dE^2 + dN^2); "
                    "|D| = |dist_F - dist_C|; Esito = OK se |D| <= Soglia."])
    ws_calc.append([])

    headers = ["#", "Tipo",
               "Punto A", "Est_A_F", "Nord_A_F",
               "Punto B", "Est_B_F", "Nord_B_F",
               "dE_F", "dN_F", "dist_F (m)",
               "Est_A_C", "Nord_A_C", "Est_B_C", "Nord_B_C",
               "dE_C", "dN_C", "dist_C (m)",
               "dist_F - dist_C", "|D| (m)", "Soglia (m)", "Esito"]
    header_row = 5
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_calc.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    for i, c in enumerate(coppie_ord, start=1):
        r = header_row + i
        pa_f = fraz_dict[c.nome_a.upper().strip()]
        pb_f = fraz_dict[c.nome_b.upper().strip()]
        pa_c = coll_dict[c.nome_a.upper().strip()]
        pb_c = coll_dict[c.nome_b.upper().strip()]

        ws_calc.cell(row=r, column=1,  value=i)
        ws_calc.cell(row=r, column=2,  value=c.tipo)
        ws_calc.cell(row=r, column=3,  value=c.nome_a)
        ws_calc.cell(row=r, column=4,  value=pa_f.est)
        ws_calc.cell(row=r, column=5,  value=pa_f.nord)
        ws_calc.cell(row=r, column=6,  value=c.nome_b)
        ws_calc.cell(row=r, column=7,  value=pb_f.est)
        ws_calc.cell(row=r, column=8,  value=pb_f.nord)
        ws_calc.cell(row=r, column=9,  value=f"=G{r}-D{r}")
        ws_calc.cell(row=r, column=10, value=f"=H{r}-E{r}")
        ws_calc.cell(row=r, column=11, value=f"=SQRT(I{r}^2+J{r}^2)")
        ws_calc.cell(row=r, column=12, value=pa_c.est)
        ws_calc.cell(row=r, column=13, value=pa_c.nord)
        ws_calc.cell(row=r, column=14, value=pb_c.est)
        ws_calc.cell(row=r, column=15, value=pb_c.nord)
        ws_calc.cell(row=r, column=16, value=f"=N{r}-L{r}")
        ws_calc.cell(row=r, column=17, value=f"=O{r}-M{r}")
        ws_calc.cell(row=r, column=18, value=f"=SQRT(P{r}^2+Q{r}^2)")
        ws_calc.cell(row=r, column=19, value=f"=R{r}-K{r}")
        ws_calc.cell(row=r, column=20, value=f"=ABS(S{r})")
        ws_calc.cell(row=r, column=21, value=c.soglia)
        ws_calc.cell(row=r, column=22, value=f'=IF(T{r}<=U{r},"OK","NOK")')

        for col_idx in (4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15,
                        16, 17, 18, 19, 20, 21):
            ws_calc.cell(row=r, column=col_idx).number_format = "0.000"
        fill = fill_nok if c.esito == "NOK" else fill_ok
        ws_calc.cell(row=r, column=22).fill = fill

    widths = [4, 8, 14, 12, 12, 14, 12, 12, 9, 9, 11,
              12, 12, 12, 12, 9, 9, 11, 14, 10, 10, 8]
    for i, w in enumerate(widths, start=1):
        ws_calc.column_dimensions[get_column_letter(i)].width = w
    ws_calc.row_dimensions[header_row].height = 30
    ws_calc.freeze_panes = f"C{header_row + 1}"

    # ----- Foglio Sintesi -----
    ws2 = wb.create_sheet("Sintesi")
    ws2.append([f"Confronto misurate - {job}"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    if comune or foglio:
        ws2.append([f"Comune: {comune or 'N.D.'}   Foglio: {foglio or 'N.D.'}"])
        ws2.append([])
    ws2.append(["Tolleranze applicate"])
    ws2["A" + str(ws2.max_row)].font = Font(bold=True)
    ws2.append(["PF - PF",                 f"{tol_pf:.2f} m"])
    ws2.append(["Dettaglio - Dettaglio",   f"{tol_det:.2f} m"])
    ws2.append(["PF - Dettaglio (mista)",  f"{tol_mista:.2f} m"])
    ws2.append([])

    ws2.append(["Tipo coppia", "Tot.", "OK", "NOK", "% conformita", "max |D| (m)"])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True)

    sint_first_data = ws2.max_row + 1
    for tipo in ("PF-PF", "PF-Det", "Det-Det"):
        sub = [c for c in coppie if c.tipo == tipo]
        if not sub:
            continue
        ok = sum(1 for c in sub if c.esito == "OK")
        max_d = max((c.delta for c in sub), default=0.0)
        pct = ok / len(sub) * 100 if sub else 0.0
        ws2.append([tipo, len(sub), ok, len(sub) - ok, round(pct, 1), round(max_d, 3)])
    sint_last_data = ws2.max_row
    if coppie:
        ok_tot = sum(1 for c in coppie if c.esito == "OK")
        max_tot = max(c.delta for c in coppie)
        pct_tot = ok_tot / len(coppie) * 100
        ws2.append(["TOTALE", len(coppie), ok_tot, len(coppie) - ok_tot,
                    round(pct_tot, 1), round(max_tot, 3)])
        ws2[ws2.max_row][0].font = Font(bold=True)

    if solo_fraz or solo_coll:
        ws2.append([])
        ws2.append(["Punti senza omologo"])
        ws2["A" + str(ws2.max_row)].font = Font(bold=True)
        if solo_fraz:
            ws2.append(["Solo nel frazionamento", ", ".join(solo_fraz)])
        if solo_coll:
            ws2.append(["Solo nel collaudo", ", ".join(solo_coll)])

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ----- Foglio Grafici -----
    _foglio_grafici(wb, coppie_ord, "Sintesi", sint_first_data, sint_last_data)

    wb.save(out_path)


def _foglio_grafici(wb, coppie_ord, sint_sheet, sint_first_data, sint_last_data):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    wsg = wb.create_sheet("Grafici")
    wsg["A1"] = "Grafici di sintesi"
    wsg["A1"].font = Font(bold=True, size=14)
    wsg["A3"] = "Coppia"
    wsg["B3"] = "|D| (m)"
    wsg["C3"] = "Soglia (m)"
    for c in (wsg["A3"], wsg["B3"], wsg["C3"]):
        c.font = Font(bold=True)

    start_row = 4
    for i, c in enumerate(coppie_ord, start=start_row):
        wsg.cell(row=i, column=1, value=f"{c.nome_a} - {c.nome_b}")
        wsg.cell(row=i, column=2, value=round(c.delta, 3))
        wsg.cell(row=i, column=3, value=round(c.soglia, 3))
    end_row = start_row + len(coppie_ord) - 1

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 11
    chart1.title = "Differenza fra mutue distanze (|D|) per coppia"
    chart1.y_axis.title = "|D| (m)"
    chart1.x_axis.title = "Coppia di punti"
    chart1.add_data(Reference(wsg, min_col=2, max_col=3, min_row=3, max_row=end_row),
                    titles_from_data=True)
    chart1.set_categories(Reference(wsg, min_col=1, max_col=1,
                                    min_row=start_row, max_row=end_row))
    chart1.height = 10
    chart1.width = 22
    wsg.add_chart(chart1, "E3")

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.style = 12
    chart2.title = "Conformita per tipo di coppia"
    chart2.y_axis.title = "Tipo di coppia"
    chart2.x_axis.title = "% conformita"
    sint = wb[sint_sheet]
    chart2.add_data(Reference(sint, min_col=5, max_col=5,
                              min_row=sint_first_data - 1, max_row=sint_last_data),
                    titles_from_data=True)
    chart2.set_categories(Reference(sint, min_col=1, max_col=1,
                                    min_row=sint_first_data, max_row=sint_last_data))
    chart2.height = 7
    chart2.width = 18
    wsg.add_chart(chart2, "E28")

    wsg.column_dimensions["A"].width = 35
    wsg.column_dimensions["B"].width = 12
    wsg.column_dimensions["C"].width = 12
