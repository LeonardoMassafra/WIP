"""
Convertitore File Coordinate -> DXF

Input accettati:
  - CSV / TXT (separatori auto-rilevati: ; , TAB spazio)
  - XLSX (prima sheet)

Struttura colonne riconosciuta (header opzionale, auto-detect):
  - Nome, Nord, Est, Quota   (convenzione topografica italiana: Y=Nord, X=Est)
  - Nome, X, Y, Z            (convenzione matematica: X=Est, Y=Nord)
  - Nome, X, Y               (2D, Z opzionale)

Header: se la 1a riga contiene keyword (nome/id, nord/est/x/y, quota/z) -> usa quei nomi.
Altrimenti fallback a Nome, X, Y, Z (con avviso).

Separatore decimale: auto-rilevato (punto o virgola).

Output DXF:
  - Layer "PUNTI"  : inserimento del blocco CROCE con nome del punto
  - Layer "NOMI"   : etichetta col nome
  - Layer "QUOTE"  : etichetta con la Z (solo se presente)

Uso:
    python coords_to_dxf.py <input.csv|.txt|.xlsx> [output.dxf]
"""

import sys
import re
import math
from pathlib import Path

import ezdxf


# =========================================================================
# Header detection
# =========================================================================

HEADER_KEYWORDS = {
    'nome': 'NOME', 'name': 'NOME', 'id': 'NOME', 'punto': 'NOME', 'pt': 'NOME',
    'n.': 'NOME', 'n': 'NOME', 'num': 'NOME', 'pid': 'NOME',
    'x': 'X', 'est': 'X', 'e': 'X', 'easting': 'X',
    'y': 'Y', 'nord': 'Y', 'n.': 'Y', 'northing': 'Y',  # attenzione: 'n.' conflitta con nome
    'z': 'Z', 'quota': 'Z', 'q': 'Z', 'elev': 'Z', 'altitude': 'Z', 'h': 'Z',
}


def _classify_header(cell):
    """Ritorna 'NOME', 'X', 'Y', 'Z' oppure None se non riconosciuto."""
    s = cell.strip().lower()
    s = s.rstrip('.:;,')
    # Parole specifiche che vincono
    if s in ('nord', 'northing'):
        return 'Y'
    if s in ('est', 'easting'):
        return 'X'
    if s in ('quota', 'elev', 'altitude'):
        return 'Z'
    if s in ('nome', 'name', 'id', 'punto', 'pt', 'num', 'pid'):
        return 'NOME'
    if s == 'x':
        return 'X'
    if s == 'y':
        return 'Y'
    if s == 'z' or s == 'h':
        return 'Z'
    if s == 'e':
        return 'X'
    if s == 'n':
        return 'Y'
    return None


def _detect_header(cells):
    """Ritorna mapping {col_index: role} se la riga e' plausibilmente un header.
    Se almeno 2 celle sono riconosciute come ruolo distinto, consideriamo header.
    """
    roles = {}
    for i, c in enumerate(cells):
        r = _classify_header(c)
        if r:
            roles[i] = r
    distinct = set(roles.values())
    if len(distinct) >= 2 and 'NOME' in distinct:
        return roles
    if len(distinct) >= 2 and {'X', 'Y'}.issubset(distinct):
        return roles
    return None


# =========================================================================
# Text / CSV parsing con auto-detect
# =========================================================================

def _sniff_separator(sample_lines):
    """Prova , ; TAB spazi. Ritorna il separatore con numero colonne consistente >=3."""
    candidates = [';', '\t', ',', None]  # None = split su whitespace (spazi/TAB)
    best = (None, 0, 0)  # (sep, consistent_count, col_count)
    for sep in candidates:
        cols = []
        for line in sample_lines:
            if sep is None:
                parts = line.split()
            else:
                parts = [p.strip() for p in line.split(sep)]
            cols.append(len(parts))
        if not cols:
            continue
        # Cerca il numero di colonne piu' frequente (>=3)
        from collections import Counter
        cnt = Counter(cols)
        most, freq = cnt.most_common(1)[0]
        if most < 3:
            continue
        if freq > best[1] or (freq == best[1] and most > best[2]):
            best = (sep, freq, most)
    return best[0], best[2]  # separatore, n_colonne


def _convert_decimal(s):
    """Converte stringa a float, gestendo virgola decimale."""
    s = s.strip().replace(' ', '')
    if not s:
        return None
    # Se ha sia '.' che ',' -> il . e' migliaia, la virgola e' decimale
    if '.' in s and ',' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s and '.' not in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def parse_text_file(path):
    """Parsa CSV/TXT restituendo (points, info)."""
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        raw = f.read()
    # normalizza newlines
    lines = [l for l in raw.splitlines() if l.strip() and not l.strip().startswith(('#', '//'))]
    if not lines:
        raise ValueError("File vuoto")

    # Prendi campione delle prime 10 righe (o meno)
    sample = lines[:10]
    sep, ncols = _sniff_separator(sample)
    if sep is None and ncols == 0:
        raise ValueError("Impossibile determinare il separatore di colonna")

    def split_row(ln):
        if sep is None:
            return ln.split()
        return [p.strip().strip('"') for p in ln.split(sep)]

    # Header detection sulla prima riga
    first_cells = split_row(lines[0])
    header_map = _detect_header(first_cells)

    points = []
    data_start = 0
    default_warning = False
    if header_map is not None:
        data_start = 1
        role_idx = {v: k for k, v in header_map.items()}  # inverso: 'X' -> col_idx
    else:
        # Nessun header: fallback default NOME, X, Y, Z
        default_warning = True
        role_idx = {'NOME': 0, 'X': 1, 'Y': 2}
        if ncols >= 4:
            role_idx['Z'] = 3

    for ln in lines[data_start:]:
        cells = split_row(ln)
        if len(cells) < 3:
            continue
        try:
            nome = cells[role_idx['NOME']].strip().strip('"')
            x = _convert_decimal(cells[role_idx['X']])
            y = _convert_decimal(cells[role_idx['Y']])
            if x is None or y is None:
                continue
            z = None
            if 'Z' in role_idx and role_idx['Z'] < len(cells):
                z = _convert_decimal(cells[role_idx['Z']])
            points.append({'nome': nome, 'X': x, 'Y': y, 'Z': z})
        except (KeyError, IndexError):
            continue

    info = {
        'separatore': repr(sep) if sep else 'spazi/TAB',
        'header_rilevato': header_map is not None,
        'avviso_default': default_warning,
        'n_colonne': ncols,
        'header_map': {v: k for k, v in role_idx.items()},  # per debug
    }
    return points, info


# =========================================================================
# Excel parsing
# =========================================================================

def parse_xlsx_file(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for row in ws.iter_rows(values_only=True):
        # ignora righe interamente vuote
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        rows.append(list(row))

    if not rows:
        raise ValueError("Foglio Excel vuoto")

    # Header detection: converti la prima riga a stringhe
    first = [str(c).strip() if c is not None else '' for c in rows[0]]
    header_map = _detect_header(first)
    default_warning = False

    if header_map is not None:
        role_idx = {v: k for k, v in header_map.items()}
        data_rows = rows[1:]
    else:
        default_warning = True
        role_idx = {'NOME': 0, 'X': 1, 'Y': 2}
        if len(rows[0]) >= 4:
            role_idx['Z'] = 3
        data_rows = rows

    points = []
    for row in data_rows:
        try:
            nome = row[role_idx['NOME']]
            nome = str(nome).strip() if nome is not None else ''
            if not nome:
                continue
            xv = row[role_idx['X']]
            yv = row[role_idx['Y']]
            x = float(xv) if isinstance(xv, (int, float)) else _convert_decimal(str(xv))
            y = float(yv) if isinstance(yv, (int, float)) else _convert_decimal(str(yv))
            if x is None or y is None:
                continue
            z = None
            if 'Z' in role_idx and role_idx['Z'] < len(row):
                zv = row[role_idx['Z']]
                if zv is not None and zv != '':
                    z = float(zv) if isinstance(zv, (int, float)) else _convert_decimal(str(zv))
            points.append({'nome': nome, 'X': x, 'Y': y, 'Z': z})
        except (KeyError, IndexError, TypeError):
            continue

    info = {
        'separatore': 'XLSX',
        'header_rilevato': header_map is not None,
        'avviso_default': default_warning,
        'n_colonne': len(rows[0]),
        'header_map': {v: k for k, v in role_idx.items()},
    }
    return points, info


# =========================================================================
# DXF writer
# =========================================================================

LAYER_COLORS = {
    'PUNTI':      2,   # giallo
    'NOMI':       7,   # bianco/nero
    'QUOTE':      3,   # verde
}


def _add_block_croce(doc):
    """Crea il blocco 'CROCE' = simbolo topografico (cerchio + + centrale).
    Lunghezza nominale = 1 unita' (scala al momento di INSERT).
    """
    block_name = 'CROCE'
    if block_name in doc.blocks:
        return block_name
    blk = doc.blocks.new(name=block_name)
    # Cerchio di raggio 0.5
    blk.add_circle((0, 0), radius=0.5, dxfattribs={'layer': '0'})
    # Crocetta interna: 2 segmentini piccoli
    blk.add_line((-0.5, 0), (0.5, 0), dxfattribs={'layer': '0'})
    blk.add_line((0, -0.5), (0, 0.5), dxfattribs={'layer': '0'})
    return block_name


def _compute_scale(points):
    """Scala del simbolo in base al bbox dei punti."""
    if not points:
        return 1.0, 1.0
    xs = [p['X'] for p in points]
    ys = [p['Y'] for p in points]
    w = max(xs) - min(xs)
    h = max(ys) - min(ys)
    diag = max(math.hypot(w, h), 1.0)
    symbol_scale = min(max(diag * 0.003, 0.1), 1.0)   # raggio cerchio ~0.1..1.0 m
    text_h = min(max(diag * 0.004, 0.15), 1.2)
    return symbol_scale, text_h


def write_dxf(points, output_path):
    doc = ezdxf.new('R2010', setup=True)
    msp = doc.modelspace()

    for name, color in LAYER_COLORS.items():
        if name not in doc.layers:
            doc.layers.add(name=name, color=color)

    block_name = _add_block_croce(doc)
    sym_scale, txt_h = _compute_scale(points)

    for p in points:
        x, y, z, nome = p['X'], p['Y'], p['Z'], p['nome']
        # INSERT del simbolo CROCE
        msp.add_blockref(
            block_name, (x, y),
            dxfattribs={'layer': 'PUNTI',
                        'xscale': sym_scale, 'yscale': sym_scale, 'zscale': sym_scale},
        )
        # Etichetta nome del punto (in alto a destra della crocetta)
        msp.add_text(
            nome, height=txt_h,
            dxfattribs={'layer': 'NOMI'},
        ).set_placement((x + sym_scale * 0.8, y + sym_scale * 0.8))
        # Etichetta quota (in basso a destra) su layer QUOTE, solo se Z presente
        if z is not None:
            msp.add_text(
                f"{z:.3f}", height=txt_h * 0.8,
                dxfattribs={'layer': 'QUOTE'},
            ).set_placement((x + sym_scale * 0.8, y - sym_scale * 1.6))

    doc.saveas(output_path)


# =========================================================================
# Main
# =========================================================================

def main(argv):
    if len(argv) < 2:
        print(__doc__)
        print("ERRORE: specifica il file di input.")
        sys.exit(1)

    input_path = Path(argv[1])
    if not input_path.exists():
        print(f"ERRORE: file non trovato: {input_path}")
        sys.exit(1)

    output_path = Path(argv[2]) if len(argv) > 2 else input_path.with_suffix('.dxf')

    print(f"Input:  {input_path.name}")
    ext = input_path.suffix.lower()
    if ext in ('.xlsx', '.xls'):
        points, info = parse_xlsx_file(str(input_path))
    else:
        points, info = parse_text_file(str(input_path))

    print(f"Formato: {ext or 'testo'}  separatore: {info['separatore']}  colonne: {info['n_colonne']}")
    if info['header_rilevato']:
        print(f"Header rilevato: {info['header_map']}")
    if info['avviso_default']:
        print("ATTENZIONE: nessun header riconosciuto -> assumo ordine Nome, X, Y, Z")
    print(f"Punti letti: {len(points)}")
    if not points:
        print("ERRORE: nessun punto valido.")
        sys.exit(1)

    # Statistiche
    xs = [p['X'] for p in points]
    ys = [p['Y'] for p in points]
    zs = [p['Z'] for p in points if p['Z'] is not None]
    print(f"Range X: [{min(xs):.3f}, {max(xs):.3f}]")
    print(f"Range Y: [{min(ys):.3f}, {max(ys):.3f}]")
    if zs:
        print(f"Range Z: [{min(zs):.3f}, {max(zs):.3f}]  ({len(zs)}/{len(points)} punti quotati)")

    write_dxf(points, str(output_path))
    print(f"Output: {output_path.name}  OK ({output_path.stat().st_size} bytes)")


if __name__ == '__main__':
    main(sys.argv)
